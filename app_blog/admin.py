from django.contrib import admin
from .models import Article, Commentaire, Formulaire, ChampFormulaire, ReponseFormulaire, ReponseChamp, FormulaireGroupe
from django.contrib.auth.models import Group
from django.http import HttpResponse
import zipfile
import os
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.utils.timezone import localtime

# Importez les bibliothèques nécessaires pour la création de fichiers Excel

class ChampFormulaireInline(admin.TabularInline):
    model = ChampFormulaire
    extra = 1
    fields = ['label', 'type_de_champ']

class FormulaireAdmin(admin.ModelAdmin):
    inlines = [ChampFormulaireInline]
    list_display = ['titre', 'description']

@admin.action(description='Exporter les réponses sélectionnées en Excel et fichiers zip')
def export_as_excel_and_zip(modeladmin, request, queryset):
    # Création d'un workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Réponses au formulaire'

    # En-têtes des colonnes
    headers = ['ID Réponse', 'Formulaire', 'Date Soumission', 'Utilisateur', 'Champ', 'Valeur']
    ws.append(headers)

    for reponse in queryset:
        reponses_champs = ReponseChamp.objects.filter(reponse_formulaire=reponse)
        for reponse_champ in reponses_champs:
            row = [
                reponse.id,
                reponse.formulaire.titre,
                reponse.date_soumission.strftime('%Y-%m-%d %H:%M'),
                reponse.utilisateur.username,
                reponse_champ.champ.label,
                reponse_champ.valeur
            ]
            ws.append(row)

    # Ajustement de la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Enregistrement du fichier Excel dans un buffer en mémoire
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Nom du fichier
    reponse_formulaire = queryset.first()
    date_completion = localtime(reponse_formulaire.date_soumission).strftime('%d_%m_%Y')
    nom_formulaire = reponse_formulaire.formulaire.titre.replace(" ", "_")
    username = reponse_formulaire.utilisateur.username
    nom_fichier = f"{date_completion}_{nom_formulaire}_{username}.zip"

    # Création du fichier ZIP
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w') as zip_file:
        # Ajout du fichier Excel au ZIP
        zip_file.writestr(f"answers_{date_completion}_{nom_formulaire}_{username}.xlsx", excel_buffer.getvalue())

        # Ajout des fichiers téléchargés au ZIP
        for reponse in queryset:
            if reponse.fichier:
                file_path = reponse.fichier.path
                zip_file.write(file_path, arcname=os.path.basename(file_path))

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={nom_fichier}'
    return response

@admin.register(ReponseFormulaire)
class ReponseFormulaireAdmin(admin.ModelAdmin):
    list_display = ['id', 'formulaire', 'date_soumission']
    actions = [export_as_excel_and_zip]

class ReponseChampAdmin(admin.ModelAdmin):
    list_display = ['reponse_formulaire', 'champ', 'valeur']

admin.site.register(Article)
admin.site.register(Commentaire)
admin.site.register(Formulaire, FormulaireAdmin)
admin.site.register(ReponseChamp, ReponseChampAdmin)

@admin.register(FormulaireGroupe)
class FormulaireGroupeAdmin(admin.ModelAdmin):
    list_display = ('formulaire', 'groupe')